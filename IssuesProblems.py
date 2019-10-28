from redminelib import Redmine
from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment
import datetime
import getpass

##Logando
print("Digite O User:")
usuario = input()
print("Digite a Senha:")
psw = getpass.getpass()
redmine = Redmine('https://redmine.querobolsa.space/', username=usuario, password=psw)


##Recebendo Valor para Intervalo de datas
print("Intervalo (yyyy-mm-dd)")
print("Digite a data de Inicio do intervalo (2019-09-12):")
data_inicio = input()
print("Digite a data de Fim do intervalo (2019-10-07):")
data_fim = input()

print("Digite o Tipo da Task: (31 - Reijeitado pela Triagem, 30 - Rejeitado na Operação")
tipoTask = int(input())


##data_inicio = datetime.datetime.strptime(data_inicio, '%d-%m-%y').strftime('%y/%m/%d')
##data_fim = datetime.datetime.strptime(data_fim, '%d-%m-%y').strftime('%y/%m/%d')

data = '><'+data_inicio+'|'+data_fim

data_example = '><2019-09-12|2019-10-07'
issues  = redmine.issue.filter(
    project_id='operacional',
    status_id= [tipoTask],
    created_on=data,
    include=['journals']
)

issuesIDs = []
issuesComplete = []

##Guardando IDs
for issue in issues:
    issuesIDs.append(issue.id)

def taskGet(task_id):
    task = redmine.issue.get(task_id, include=['journals'])
    print(task.author['name'])
    for x in task.journals:
        dataTask = task.created_on.split("T")[0]
        #print(x.details)
        for z in x.details:
            array = []
            if(z['name'] == 'status_id' and z['new_value'] == '30'):
                if hasattr(task, 'assigned_to'):
                    issuesComplete.append([task.id, 'Reijeitado Site Ops', x['notes'],task.assigned_to['name'],task.author['name'], dataTask])
                else:
                    issuesComplete.append([task.id, 'Reijeitado Site Ops', x['notes'],'_________',task.author['name'], dataTask])
            elif(z['name'] == 'status_id' and z['new_value'] == '31'):
                if hasattr(task, 'assigned_to'):
                    issuesComplete.append([task.id, 'Reijeitado Site Ops Triagem', x['notes'],task.assigned_to['name'],task.author['name'], dataTask])
                else:
                    issuesComplete.append([task.id, 'Reijeitado Site Ops Triagem', x['notes'],'_________',task.author['name'], dataTask])


print("Carregando dados:")
for z in issuesIDs:
    print(z)
    taskGet(z)

##Nomeando Titulos para diferentes métodos
if(tipoTask == 31):
    tipo = 'PUSH_Triagem_'
else:
    tipo = 'PUSH_Operação_'

    
##Gravando no Excel
wb = Workbook()
fileText = './data/'+tipo+str(datetime.datetime.today()).split(".")[0].replace(' ', '_').replace(':', '.')+'.xlsx'

ws1 = wb.active
ws1.title = "Tasks - Reijeitadas - Site Ops"
_ = ws1.cell(column=1, row=1, value="Task_id")
_ = ws1.cell(column=2, row=1, value="Tipo")
_ = ws1.cell(column=3, row=1, value="Notas")
_ = ws1.cell(column=4, row=1, value="Atribuido Para")
_ = ws1.cell(column=5, row=1, value="Autor")
_ = ws1.cell(column=6, row=1, value="Data")


##Colocando Estilo nos cabeçalhos
ft = Font(name='Calibri',size=11,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')
ws1['A1'].font = ft
ws1['B1'].font = ft
ws1['C1'].font = ft
ws1['D1'].font = ft
ws1['E1'].font = ft
ws1['F1'].font = ft
##Registrando dados no Excel e colocando o tamanho das colunas
maxLenght = [0, 0, 0, 0, 0, 0]
for row in range(2,len(issuesComplete)+1):
    for col in range(1, 7):
        valor = len(str(issuesComplete[row-2][col-1]))
        if(maxLenght[col-1] < valor):
            maxLenght[col-1] = valor
        _ = ws1.cell(column=col, row=row, value=issuesComplete[row-2][col-1])
ws1.column_dimensions['A'].width = maxLenght[0]+10
ws1.column_dimensions['B'].width = maxLenght[1]
ws1.column_dimensions['C'].width = maxLenght[2]
ws1.column_dimensions['D'].width = maxLenght[3]
ws1.column_dimensions['E'].width = maxLenght[4]
ws1.column_dimensions['F'].width = maxLenght[5]


##Salvando arquivo
wb.save(filename = fileText)

input("Finalizado com sucesso... Aperte Enter para finalizar")
